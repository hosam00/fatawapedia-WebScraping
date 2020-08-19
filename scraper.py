from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from time import sleep

# full page numbers 35760
number = 31820
driver = webdriver.Chrome(executable_path="E:/test/chromedriver.exe")
driver.maximize_window()
row = 31781
while number <= 35760:
    url = f'http://fatawapedia.com/558-%D8%A7%D9%84%D8%AA%D8%B5%D9%86%D9%8A%D9%81-%D8%A7%D9%84%D9%81%D9%82%D9%87%D9%8A/page/{number}'
    driver.get(url)
    # sleep(3)
    # links = [l.get_attribute('href') for l in WebDriverWait(driver, 1).until(EC.visibility_of_all_elements_located((By.XPATH, '//*[@id="content_cat"]/div/a')))]
    # links = [l.get_attribute('href') for l in driver.find_elements_by_xpath('//*[@id="content_cat"]/div/a')]
    links = [l.get_attribute('href') for l in driver.find_elements_by_css_selector('#content_cat > div > a')]
    wb = load_workbook('sample.xlsx')
    sheet = wb.active
    for link in links:
        driver.get(link)
        try:
            category = driver.find_element_by_xpath('//*[@id="content_item"]/div[1]/a[1]').text
        except:
            category = ""
        try:
            cat = driver.find_element_by_xpath('//*[@id="content_item"]/div[1]/a[2]').text
        except:
            cat = ""
        try:
            name = driver.find_element_by_xpath('//*[@id="content_item"]/div[2]/div[2]/div[1]/h1').text
        except:
            name = ""
        try:
            question = driver.find_element_by_xpath('//*[@id="content_item"]/div[2]/div[2]/div[2]/p[2]').text
        except:
            question = ""
        try:
            answer = driver.find_element_by_xpath('//*[@id="content_item"]/div[2]/div[2]/div[3]/p[2]').text
        except:
            answer = ""
        try:
            source = driver.find_element_by_xpath('//*[@id="content_item"]/div[2]/div[2]/div[4]/p[2]').text
        except:
            source = ""
        sheet.cell(row=row, column=1, value=name)
        sheet.cell(row=row, column=2, value=category)
        sheet.cell(row=row, column=3, value=cat)
        sheet.cell(row=row, column=4, value=question)
        sheet.cell(row=row, column=5, value=answer)
        sheet.cell(row=row, column=6, value=source)
        sheet.cell(row=row, column=7, value=link)
        print(row)
        row += 1
    print(f'Page: {number}')
    wb.save('sample.xlsx')
    number += 20
driver.quit()
